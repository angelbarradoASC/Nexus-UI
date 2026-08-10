"""
scripts/tests_to_excel.py
--------------------------
Corre la suite de pytest completa (o la ruta que le pases) y vuelca el
resultado a un .xlsx para auditar fuera de la terminal: una fila por test,
con fichero, clase, nombre, resultado, duracion y el mensaje de error si
fallo.

Uso:
    .venv/Scripts/python.exe scripts/tests_to_excel.py
    .venv/Scripts/python.exe scripts/tests_to_excel.py tests/unit
    .venv/Scripts/python.exe scripts/tests_to_excel.py --out mi_reporte.xlsx

No requiere pytest-json-report ni pandas — usa --junitxml (ya viene con
pytest) + openpyxl (unica dependencia nueva, solo de test-tooling).
"""

from __future__ import annotations

import argparse
import subprocess
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parent.parent
_STATUS_FILL = {
    "passed":  PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "failed":  PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "error":   PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "skipped": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
}


def run_pytest(target: str, junit_path: Path) -> int:
    cmd = [
        sys.executable, "-m", "pytest", target,
        f"--junitxml={junit_path}",
        "-q",
    ]
    print(f"Ejecutando: {' '.join(cmd)}")
    result = subprocess.run(cmd, cwd=REPO_ROOT)
    return result.returncode


def parse_junit(junit_path: Path) -> list[dict]:
    tree = ET.parse(junit_path)
    root = tree.getroot()
    suites = [root] if root.tag == "testsuite" else list(root.findall("testsuite"))

    rows = []
    for suite in suites:
        for case in suite.findall("testcase"):
            classname = case.get("classname", "")
            name = case.get("name", "")
            duration = float(case.get("time", "0") or 0)

            status = "passed"
            message = ""
            failure = case.find("failure")
            error = case.find("error")
            skipped = case.find("skipped")
            if failure is not None:
                status = "failed"
                message = (failure.get("message") or failure.text or "").strip()
            elif error is not None:
                status = "error"
                message = (error.get("message") or error.text or "").strip()
            elif skipped is not None:
                status = "skipped"
                message = (skipped.get("message") or "").strip()

            # classname suele ser "tests.unit.test_foo.TestBar" — separar
            # el path de fichero del nombre de clase para que se pueda
            # filtrar/agrupar por fichero en Excel.
            parts = classname.split(".")
            file_path = "/".join(parts[:-1]) + ".py" if len(parts) > 1 else classname
            cls = parts[-1] if len(parts) > 1 else ""

            rows.append({
                "file": file_path,
                "class": cls,
                "test": name,
                "status": status,
                "duration_s": round(duration, 3),
                "message": message[:500],
            })
    return rows


def write_excel(rows: list[dict], out_path: Path) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Tests"
    headers = ["Fichero", "Clase", "Test", "Resultado", "Duracion (s)", "Mensaje"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([
            row["file"], row["class"], row["test"],
            row["status"], row["duration_s"], row["message"],
        ])
        fill = _STATUS_FILL.get(row["status"])
        if fill:
            ws.cell(row=ws.max_row, column=4).fill = fill

    widths = [45, 30, 45, 12, 12, 70]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"

    # ── Hoja resumen por fichero ──
    summary = wb.create_sheet("Resumen por fichero")
    summary.append(["Fichero", "Passed", "Failed", "Error", "Skipped", "Total"])
    for cell in summary[1]:
        cell.font = Font(bold=True)

    by_file: dict[str, dict[str, int]] = {}
    for row in rows:
        bucket = by_file.setdefault(row["file"], {"passed": 0, "failed": 0, "error": 0, "skipped": 0})
        bucket[row["status"]] += 1

    for file_path, counts in sorted(by_file.items()):
        total = sum(counts.values())
        summary.append([
            file_path, counts["passed"], counts["failed"], counts["error"], counts["skipped"], total,
        ])
    for i, width in enumerate([50, 10, 10, 10, 10, 10], start=1):
        summary.column_dimensions[get_column_letter(i)].width = width
    summary.freeze_panes = "A2"

    wb.save(out_path)
    print(f"Excel escrito en: {out_path}")

    total = len(rows)
    passed = sum(1 for r in rows if r["status"] == "passed")
    failed = sum(1 for r in rows if r["status"] in ("failed", "error"))
    skipped = sum(1 for r in rows if r["status"] == "skipped")
    print(f"Total: {total} | passed: {passed} | failed/error: {failed} | skipped: {skipped}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("target", nargs="?", default="tests", help="Ruta de tests a correr (default: tests)")
    parser.add_argument("--out", default="test_report.xlsx", help="Fichero .xlsx de salida")
    parser.add_argument("--junit-xml", default=None, help="Reusar un junit.xml ya generado en vez de correr pytest")
    args = parser.parse_args()

    out_path = (REPO_ROOT / args.out) if not Path(args.out).is_absolute() else Path(args.out)

    if args.junit_xml:
        junit_path = Path(args.junit_xml)
    else:
        junit_path = REPO_ROOT / "_junit_report.xml"
        run_pytest(args.target, junit_path)

    rows = parse_junit(junit_path)
    write_excel(rows, out_path)


if __name__ == "__main__":
    main()
